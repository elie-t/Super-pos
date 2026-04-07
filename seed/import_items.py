"""
Import items, categories, and suppliers from the Excel file into the local SQLite DB.

Key rule: rows with the same Title are treated as the SAME item.
Every distinct Code for that title becomes an extra barcode so scanning
any of them finds the item at the POS.

Usage:
    python seed/import_items.py [--file PATH] [--clear] [--batch N]

Options:
    --file   Path to Excel file (default: ~/Downloads/items_list.xlsx)
    --clear  Delete all existing items/barcodes/prices/stock before import
    --batch  Commit batch size (default: 500)
"""
import argparse
import sys
import os
import uuid
from pathlib import Path
from collections import defaultdict

# Allow running from project root
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy.orm import Session
from database.engine import init_db, get_session
from database.models.items import Item, ItemBarcode, ItemPrice, Category, Warehouse, ItemStock
from database.models.parties import Supplier
from config import LBP_PRICE_THRESHOLD


def clean_str(raw) -> str:
    if not raw:
        return ""
    return str(raw).strip().replace("\xa0", "").strip()


def detect_currency(price: float) -> str:
    return "LBP" if price >= LBP_PRICE_THRESHOLD else "USD"


def get_or_create_category(session: Session, name: str, cache: dict) -> Category:
    if name in cache:
        return cache[name]
    cat = session.query(Category).filter_by(name=name).first()
    if not cat:
        cat = Category(id=str(uuid.uuid4()), name=name)
        session.add(cat)
        session.flush()
    cache[name] = cat
    return cat


def get_or_create_supplier(session: Session, name: str, cache: dict) -> Supplier:
    if name in cache:
        return cache[name]
    sup = session.query(Supplier).filter_by(name=name).first()
    if not sup:
        sup = Supplier(id=str(uuid.uuid4()), name=name)
        session.add(sup)
        session.flush()
    cache[name] = sup
    return sup


def get_default_warehouse(session: Session) -> Warehouse:
    wh = session.query(Warehouse).filter_by(is_default=True).first()
    if not wh:
        raise RuntimeError("No default warehouse found. Run seed/sample_data.py first.")
    return wh


def clear_items(session: Session):
    print("Clearing existing items, barcodes, prices, stock…")
    from sqlalchemy import text
    # Disable FK so we can delete items without cascading into invoices/movements
    session.execute(text("PRAGMA foreign_keys=OFF"))
    session.query(ItemBarcode).delete()
    session.query(ItemPrice).delete()
    session.query(ItemStock).delete()
    session.query(Item).delete()
    session.commit()
    session.execute(text("PRAGMA foreign_keys=ON"))
    print("  Done.\n")


def load_xlsx(xlsx_path: str):
    """Read all data rows and return list of dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    iter_rows = ws.iter_rows(values_only=True)
    header = next(iter_rows)
    rows = list(iter_rows)
    wb.close()
    return rows


def import_items(xlsx_path: str, batch_size: int = 500, do_clear: bool = False) -> None:
    print(f"Opening: {xlsx_path}")
    raw_rows = load_xlsx(xlsx_path)
    print(f"  Total rows in file: {len(raw_rows):,}")

    init_db()
    session: Session = get_session()

    if do_clear:
        clear_items(session)

    warehouse = get_default_warehouse(session)

    # ── Group rows by normalised title ──────────────────────────────────────
    # Each group collects all codes, barcodes, and picks the canonical values
    # from the first row (cost, price, supplier, category, stock, flags).

    # title_key → { 'rows': [raw_row, …] }
    groups: dict[str, list] = defaultdict(list)

    for row in raw_rows:
        _, code, barcode_raw, pkg, title, *_ = row
        if not code or not title:
            continue
        key = clean_str(title).upper()
        groups[key].append(row)

    print(f"  Unique item titles  : {len(groups):,}\n")

    cat_cache: dict = {}
    sup_cache: dict = {}

    # Pre-load existing titles and codes into memory for upsert
    existing_by_title: dict[str, str] = {}  # normalised title → item.id
    for item in session.query(Item).all():
        existing_by_title[item.name.upper()] = item.id

    existing_barcodes: set = {
        b.barcode for b in session.query(ItemBarcode.barcode).all()
    }

    inserted = 0
    updated  = 0
    skipped  = 0
    total    = 0

    try:
        for title_key, rows in groups.items():
            total += 1

            # ── Pick canonical values from first row ─────────────────────
            first = rows[0]
            _, code0, barcode0_raw, pkg0, title_raw, stock_raw, cost_raw, \
                sup_raw, price_raw, subgroup_raw, brand_raw, \
                visible_raw, active_raw, featured_raw = first

            title_str     = clean_str(title_raw)
            code_str      = str(int(code0)) if isinstance(code0, float) else clean_str(code0)
            pkg_int       = int(pkg0) if pkg0 else 1
            cost_val      = float(cost_raw)  if cost_raw  else 0.0
            price_val     = float(price_raw) if price_raw else 0.0
            stock_val     = float(stock_raw) if stock_raw else 0.0
            subgroup_str  = clean_str(subgroup_raw) or "GENERAL"
            sup_name      = clean_str(sup_raw) or "CASH PURCHASE"
            is_active_flag   = bool(active_raw)   if active_raw   is not None else True
            is_visible_flag  = bool(visible_raw)  if visible_raw  is not None else True
            is_featured_flag = bool(featured_raw) if featured_raw is not None else False

            price_currency = detect_currency(price_val)
            cost_currency  = "USD"

            category = get_or_create_category(session, subgroup_str, cat_cache)
            supplier = get_or_create_supplier(session, sup_name,     sup_cache)

            # ── Collect ALL unique codes and barcodes from every row ─────
            all_codes:    list[tuple[str, int]] = []  # (code_str, pack_qty)
            all_barcodes: list[tuple[str, int]] = []  # (barcode_str, pack_qty)
            seen_codes    : set[str] = set()
            seen_barcodes : set[str] = set()

            for row in rows:
                _, rc, rb, rp, *_ = row
                rp_int = int(rp) if rp else 1

                rc_str = str(int(rc)) if isinstance(rc, float) else clean_str(rc)
                if rc_str and rc_str not in seen_codes:
                    seen_codes.add(rc_str)
                    all_codes.append((rc_str, rp_int))

                rb_str = clean_str(rb)
                if rb_str and rb_str not in seen_barcodes:
                    seen_barcodes.add(rb_str)
                    all_barcodes.append((rb_str, rp_int))

            # ── Upsert item ──────────────────────────────────────────────
            if title_key in existing_by_title:
                item_id = existing_by_title[title_key]
                item = session.query(Item).filter_by(id=item_id).first()
                if not item:
                    skipped += 1
                    continue
                # Update modifiable fields
                item.category_id  = category.id
                item.default_supplier_id = supplier.id
                item.cost_price   = cost_val
                item.cost_currency = cost_currency
                item.is_active    = True
                item.is_pos_featured = is_active_flag
                item.is_online    = is_active_flag
                item.is_visible   = is_visible_flag
                item.is_featured  = is_featured_flag

                # Update prices for existing items
                if price_val > 0:
                    for ptype in ("retail", "individual"):
                        existing_price = session.query(ItemPrice).filter_by(
                            item_id=item.id, price_type=ptype, currency=price_currency
                        ).first()
                        if existing_price:
                            existing_price.amount = price_val
                        else:
                            session.add(ItemPrice(
                                id=str(uuid.uuid4()),
                                item_id=item.id,
                                price_type=ptype,
                                amount=price_val,
                                currency=price_currency,
                                is_default=True,
                            ))

                session.flush()
                updated += 1
            else:
                item = Item(
                    id=str(uuid.uuid4()),
                    code=code_str,
                    name=title_str,
                    category_id=category.id,
                    default_supplier_id=supplier.id,
                    unit="PCS",
                    pack_size=pkg_int,
                    cost_price=cost_val,
                    cost_currency=cost_currency,
                    vat_rate=0.0,
                    is_active=True,
                    is_pos_featured=is_active_flag,
                    is_online=is_active_flag,
                    is_visible=is_visible_flag,
                    is_featured=is_featured_flag,
                )
                session.add(item)
                session.flush()
                existing_by_title[title_key] = item.id
                inserted += 1

                # Prices (retail + individual so POS works)
                if price_val > 0:
                    for ptype in ("retail", "individual"):
                        session.add(ItemPrice(
                            id=str(uuid.uuid4()),
                            item_id=item.id,
                            price_type=ptype,
                            amount=price_val,
                            currency=price_currency,
                            is_default=True,
                        ))

                # Opening stock
                session.add(ItemStock(
                    id=str(uuid.uuid4()),
                    item_id=item.id,
                    warehouse_id=warehouse.id,
                    quantity=stock_val,
                ))

            # ── Barcodes: all codes + all barcodes from every row ────────
            # Codes first — each code is also a scannable identifier
            for bc_val, bc_pkg in all_codes + all_barcodes:
                if bc_val in existing_barcodes:
                    continue
                existing_barcodes.add(bc_val)
                is_primary = (bc_val == all_barcodes[0][0]) if all_barcodes else (bc_val == all_codes[0][0])
                session.add(ItemBarcode(
                    id=str(uuid.uuid4()),
                    item_id=item.id,
                    barcode=bc_val,
                    is_primary=is_primary,
                    pack_qty=bc_pkg,
                ))

            # ── Batch commit ─────────────────────────────────────────────
            if total % batch_size == 0:
                session.commit()
                print(f"  {total:,} / {len(groups):,} | inserted {inserted:,} | updated {updated:,} | skipped {skipped:,}")

        session.commit()
        print(f"\n✓ Import complete.")
        print(f"  Items inserted  : {inserted:,}")
        print(f"  Items updated   : {updated:,}")
        print(f"  Items skipped   : {skipped:,}")
        print(f"  Categories      : {len(cat_cache)}")
        print(f"  Suppliers       : {len(sup_cache)}")

    except Exception as exc:
        session.rollback()
        print(f"\n✗ Import failed at item #{total}: {exc}")
        raise
    finally:
        session.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import items from Excel into TannouryMarket POS")
    parser.add_argument("--file",  default=str(Path.home() / "Downloads/items_list.xlsx"))
    parser.add_argument("--clear", action="store_true", help="Wipe existing items before import")
    parser.add_argument("--batch", type=int, default=500)
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"File not found: {args.file}")
        sys.exit(1)

    import_items(args.file, args.batch, do_clear=args.clear)
